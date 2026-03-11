from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from fractions import Fraction
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class CollatzRow:
    n: int
    me_value: float | int
    path: str
    ops: str
    div2_count: int
    mul3p1_count: int


def collatz_trace(n: int) -> Tuple[List[int], str, int, int]:
    """
    Returns (values, ops, div2_count, mul3p1_count).
    values includes the starting n and ends when the cycle 1->4->2->1 appears (inclusive).
    ops is a string where:
      - 'e' means divide by 2 (even step)
      - 'm' means 3n+1 (odd step)
    """
    if n <= 0:
        raise ValueError("n must be a positive integer")

    values = [n]
    ops_chars: List[str] = []
    div2_count = 0
    mul3p1_count = 0

    x = n
    while True:
        if x % 2 == 0:
            x = x // 2
            ops_chars.append("e")
            div2_count += 1
        else:
            x = 3 * x + 1
            ops_chars.append("m")
            mul3p1_count += 1

        values.append(x)
        if len(values) >= 4 and values[-4:] == [1, 4, 2, 1]:
            break

    return values, "".join(ops_chars), div2_count, mul3p1_count


def format_path(values: List[int]) -> str:
    return "->".join(str(v) for v in values)


def format_ops(ops: str) -> str:
    """
    Compress consecutive 'e' (divide-by-2 ops) as 'ne', e.g. 'eeemem' -> '3emem'.
    'm' tokens are kept as-is.
    """
    if not ops:
        return ops

    out: List[str] = []
    i = 0
    while i < len(ops):
        ch = ops[i]
        if ch == "e":
            j = i
            while j < len(ops) and ops[j] == "e":
                j += 1
            run = j - i
            out.append("e" if run == 1 else f"{run}e")
            i = j
            continue
        out.append(ch)
        i += 1
    return "".join(out)


def apply_m_then_e(n: int, m_count: int, e_count: int) -> float | int:
    x = n
    for _ in range(m_count):
        x = 3 * x + 1

    if e_count <= 0:
        return x

    value = Fraction(x, 2**e_count)
    if value.denominator == 1:
        return int(value)
    return float(value)


def build_rows(start: int, end: int) -> List[CollatzRow]:
    rows: List[CollatzRow] = []
    for n in range(start, end + 1):
        values, ops, div2_count, mul3p1_count = collatz_trace(n)
        rows.append(
            CollatzRow(
                n=n,
                me_value=apply_m_then_e(n, mul3p1_count, div2_count),
                path=format_path(values),
                ops=format_ops(ops),
                div2_count=div2_count,
                mul3p1_count=mul3p1_count,
            )
        )
    return rows


def autosize_columns(ws) -> None:
    # Keep it simple: set widths based on max string length per column, with caps.
    max_widths = {}
    for row in ws.iter_rows(values_only=True):
        for col_idx, cell_value in enumerate(row, start=1):
            text = "" if cell_value is None else str(cell_value)
            max_widths[col_idx] = max(max_widths.get(col_idx, 0), len(text))

    for col_idx, width in max_widths.items():
        col_letter = get_column_letter(col_idx)
        # Reasonable caps for readability in Excel.
        if col_idx == 2:
            ws.column_dimensions[col_letter].width = min(max(12, width + 2), 120)
        else:
            ws.column_dimensions[col_letter].width = min(max(10, width + 2), 60)


def write_excel(rows: List[CollatzRow], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collatz"

    ws.append(["数字", "中间计算结果", "操作串", "除以2次数(e)", "乘以3加1次数(m)", "先m后e结果"])
    for r in rows:
        ws.append([r.n, r.path, r.ops, r.div2_count, r.mul3p1_count, r.me_value])

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    wb.save(out_path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Export Collatz traces to Excel.")
    p.add_argument("--start", type=int, default=1, help="start integer (inclusive)")
    p.add_argument("--end", type=int, default=1000000, help="end integer (inclusive)")
    p.add_argument(
        "--out",
        type=str,
        default="collatz_1_1000000.xlsx",
        help="output .xlsx filename (will be written under output/)",
    )
    return p.parse_args()


def resolve_out_path(out_arg: str) -> str:
    os.makedirs("output", exist_ok=True)
    return os.path.join("output", os.path.basename(out_arg))


def main() -> None:
    args = parse_args()
    if args.start <= 0 or args.end <= 0:
        raise SystemExit("start/end must be positive integers")
    if args.start > args.end:
        raise SystemExit("start must be <= end")

    rows = build_rows(args.start, args.end)
    out_path = resolve_out_path(args.out)
    write_excel(rows, out_path)
    print(f"Done. Wrote {len(rows)} rows to: {out_path}")


if __name__ == "__main__":
    main()

